# reimbursement/admin.py
import os
import zipfile
from io import BytesIO
from datetime import datetime
from django.contrib import admin
from django.http import HttpResponse, FileResponse
from django.utils.html import format_html
from django.urls import path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from .models import ReimbursementRequest, Notice, AccountBook
from .admin_site import restricted_admin_site

@admin.register(ReimbursementRequest, site=restricted_admin_site)
class ReimbursementRequestAdmin(admin.ModelAdmin):
    list_display = ('submission_date', 'real_name', 'reason', 'amount', 'is_taxi_invoice', 'status', 'user', 'download_link', 'itinerary_download_link', 'pdf_file_link')
    list_filter = ('status', 'is_taxi_invoice', 'submission_date')
    search_fields = ('real_name', 'reason', 'user__username')
    readonly_fields = ('user', 'submission_date', 'last_modified_date', 'download_link', 'itinerary_download_link', 'pdf_file_link', 'itinerary_file_link')
    actions = ['download_approved_invoices', 'delete_unapproved_requests', 'export_approved_to_excel', 'delete_pdf_files']
    fieldsets = (
        ('申请详情', {'fields': ('user', 'real_name', 'reason', 'amount', 'invoice_pdf', 'is_taxi_invoice', 'itinerary_pdf', 'remarks')}),
        ('审核区域', {'fields': ('status', 'rejection_reason')}),
        ('日期信息', {'fields': ('submission_date', 'last_modified_date')}),
        ('文件管理', {'fields': ('download_link', 'itinerary_download_link', 'pdf_file_link', 'itinerary_file_link')}),
    )
    
    def download_link(self, obj):
        """显示发票下载链接"""
        if obj.invoice_pdf and obj.status == 'approved':
            url = obj.invoice_pdf.url
            filename = os.path.basename(obj.invoice_pdf.name)
            return format_html('<a href="{}" download="{}" target="_blank">📥 下载发票PDF</a>', url, filename)
        elif obj.invoice_pdf:
            return '<span style="color: #999;">⏳ 待审核通过后可下载</span>'
        return '<span style="color: #ccc;">无文件</span>'
    download_link.short_description = '发票下载'
    
    def itinerary_download_link(self, obj):
        """显示行程单下载链接"""
        if obj.itinerary_pdf and obj.status == 'approved':
            url = obj.itinerary_pdf.url
            filename = os.path.basename(obj.itinerary_pdf.name)
            return format_html('<a href="{}" download="{}" target="_blank">📥 下载行程单</a>', url, filename)
        elif obj.itinerary_pdf:
            return '<span style="color: #999;">⏳ 待审核通过后可下载</span>'
        return '<span style="color: #ccc;">无行程单</span>'
    itinerary_download_link.short_description = '行程单下载'
    
    def download_approved_invoices(self, request, queryset):
        """批量下载已审核通过的发票（打包成ZIP）"""
        approved = queryset.filter(status='approved', invoice_pdf__isnull=False)
        
        if not approved.exists():
            self.message_user(request, '所选申请中没有已审核通过的发票可下载', level='warning')
            return
        
        # 创建ZIP文件
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for req in approved:
                if req.invoice_pdf:
                    try:
                        file_path = req.invoice_pdf.path
                        file_name = os.path.basename(req.invoice_pdf.name)
                        zip_file.write(file_path, arcname=file_name)
                    except Exception as e:
                        self.message_user(request, f'文件 {req.invoice_pdf.name} 读取失败: {str(e)}', level='error')
        
        # 返回ZIP文件
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer.read(), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="approved_invoices_{queryset.count()}files.zip"'
        
        self.message_user(request, f'成功打包 {approved.count()} 个已审核发票')
        return response
    
    download_approved_invoices.short_description = '📦 批量下载已审核通过的发票（ZIP）'
    
    def delete_unapproved_requests(self, request, queryset):
        """删除未审核通过的报销申请（包括文件）"""
        # 只能删除待审核或审核不通过的申请
        unapproved = queryset.exclude(status='approved')
        
        if not unapproved.exists():
            self.message_user(request, '所选申请都已审核通过，无法删除！', level='warning')
            return
        
        deleted_count = 0
        for req in unapproved:
            # 删除关联的文件
            if req.invoice_pdf:
                try:
                    if os.path.exists(req.invoice_pdf.path):
                        os.remove(req.invoice_pdf.path)
                except Exception as e:
                    self.message_user(request, f'发票文件删除失败: {str(e)}', level='error')
            if req.itinerary_pdf:
                try:
                    if os.path.exists(req.itinerary_pdf.path):
                        os.remove(req.itinerary_pdf.path)
                except Exception as e:
                    self.message_user(request, f'行程单文件删除失败: {str(e)}', level='error')
            # 删除数据库记录
            req.delete()
            deleted_count += 1
        
        self.message_user(request, f'成功删除 {deleted_count} 条未审核通过的申请及其文件')
    
    delete_unapproved_requests.short_description = '🗑️ 删除未审核通过的申请（释放空间）'
    
    def pdf_file_link(self, obj):
        """显示PDF文件信息和删除按钮"""
        if obj.invoice_pdf:
            filename = os.path.basename(obj.invoice_pdf.name)
            file_size = ''
            try:
                size_bytes = obj.invoice_pdf.size
                if size_bytes < 1024:
                    file_size = f'{size_bytes} B'
                elif size_bytes < 1024 * 1024:
                    file_size = f'{size_bytes / 1024:.1f} KB'
                else:
                    file_size = f'{size_bytes / (1024 * 1024):.1f} MB'
            except:
                file_size = '未知大小'
            
            return format_html(
                '<div style="padding: 8px; background: #f8f9fa; border-radius: 4px;">'
                '<div style="margin-bottom: 4px;"><strong>📄 {}</strong></div>'
                '<div style="color: #666; font-size: 12px;">文件大小: {}</div>'
                '</div>',
                filename, file_size
            )
        return format_html('<span style="color: #999;">无PDF文件</span>')
    
    pdf_file_link.short_description = 'PDF文件信息'
    
    def itinerary_file_link(self, obj):
        """显示行程单文件信息"""
        if obj.itinerary_pdf:
            filename = os.path.basename(obj.itinerary_pdf.name)
            file_size = ''
            try:
                size_bytes = obj.itinerary_pdf.size
                if size_bytes < 1024:
                    file_size = f'{size_bytes} B'
                elif size_bytes < 1024 * 1024:
                    file_size = f'{size_bytes / 1024:.1f} KB'
                else:
                    file_size = f'{size_bytes / (1024 * 1024):.1f} MB'
            except:
                file_size = '未知大小'
            
            return format_html(
                '<div style="padding: 8px; background: #f8f9fa; border-radius: 4px;">'
                '<div style="margin-bottom: 4px;"><strong>📄 {}</strong></div>'
                '<div style="color: #666; font-size: 12px;">文件大小: {}</div>'
                '</div>',
                filename, file_size
            )
        return format_html('<span style="color: #999;">无行程单文件</span>')
    
    itinerary_file_link.short_description = '行程单文件信息'
    
    def delete_pdf_files(self, request, queryset):
        """批量删除选中申请的PDF文件"""
        has_pdf = queryset.exclude(invoice_pdf='').exclude(invoice_pdf__isnull=True)
        
        if not has_pdf.exists():
            self.message_user(request, '所选申请中没有PDF文件', level='warning')
            return
        
        deleted_count = 0
        for req in has_pdf:
            if req.invoice_pdf:
                try:
                    # 删除物理文件
                    if os.path.exists(req.invoice_pdf.path):
                        os.remove(req.invoice_pdf.path)
                    # 清空数据库字段
                    req.invoice_pdf = None
                    req.save()
                    deleted_count += 1
                except Exception as e:
                    self.message_user(request, f'删除 {req.real_name} 的PDF文件失败: {str(e)}', level='error')
        
        self.message_user(request, f'成功删除 {deleted_count} 个PDF文件（申请记录保留）')
    
    delete_pdf_files.short_description = '🗑️ 删除选中申请的PDF文件'
    
    def export_approved_to_excel(self, request, queryset):
        """导出已审核通过的报销申请到Excel表格"""
        approved = queryset.filter(status='approved').order_by('submission_date')
        
        if not approved.exists():
            self.message_user(request, '所选申请中没有已审核通过的记录', level='warning')
            return
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "已审核报销申请"
        
        # 设置表头
        headers = ['提交日期', '提交人姓名', '报销事由', '金额', '备注']
        ws.append(headers)
        
        # 设置表头样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 添加数据
        for req in approved:
            ws.append([
                req.submission_date.strftime('%Y-%m-%d'),
                req.real_name,
                req.reason,
                -float(req.amount),  # 负数形式，方便计算
                req.remarks or ''
            ])
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 40
        
        # 设置金额列格式为标准数字格式（带两位小数）
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=4)
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right')
        
        # 保存到内存
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # 返回Excel文件
        filename = f'approved_reimbursements_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = HttpResponse(
            excel_buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        self.message_user(request, f'成功导出 {approved.count()} 条已审核记录到Excel')
        return response
    
    export_approved_to_excel.short_description = '📊 导出已审核通过的申请到Excel表格'
    
@admin.register(Notice, site=restricted_admin_site)
class NoticeAdmin(admin.ModelAdmin):
    list_display = ('title', 'is_active', 'priority', 'created_at', 'updated_at')
    list_filter = ('is_active', 'created_at')
    search_fields = ('title', 'content')
    list_editable = ('is_active', 'priority')
    fieldsets = (
        ('基本信息', {'fields': ('title', 'content')}),
        ('显示设置', {'fields': ('is_active', 'priority')}),
        ('时间信息', {'fields': ('created_at', 'updated_at')}),
    )
    readonly_fields = ('created_at', 'updated_at')
    ordering = ['-priority', '-created_at']

@admin.register(AccountBook, site=restricted_admin_site)
class AccountBookAdmin(admin.ModelAdmin):
    list_display = ('entry_date', 'real_name', 'reason', 'amount_display', 'entry_type', 'remarks_short', 'reimbursement_link')
    list_filter = ('entry_type', 'entry_date')
    search_fields = ('real_name', 'reason', 'remarks')
    date_hierarchy = 'entry_date'
    actions = ['export_to_excel', 'calculate_balance']
    
    fieldsets = (
        ('记账信息', {
            'fields': ('entry_date', 'entry_type', 'real_name', 'reason', 'amount', 'remarks')
        }),
        ('关联信息', {
            'fields': ('reimbursement', 'created_by', 'created_at'),
            'classes': ('collapse',)
        }),
    )
    
    readonly_fields = ('created_at', 'created_by')
    
    def get_readonly_fields(self, request, obj=None):
        """如果是从报销自动生成的，大部分字段只读"""
        if obj and obj.reimbursement:
            return self.readonly_fields + ('entry_date', 'real_name', 'reason', 'amount', 'reimbursement')
        return self.readonly_fields
    
    def save_model(self, request, obj, form, change):
        """保存时记录创建人"""
        if not change:  # 新建时
            obj.created_by = request.user
        super().save_model(request, obj, form, change)
    
    def amount_display(self, obj):
        """格式化金额显示"""
        if obj.entry_type == 'income':
            return format_html('<span style="color: green;">+¥{}</span>', f'{obj.amount:.2f}')
        else:
            return format_html('<span style="color: red;">-¥{}</span>', f'{obj.amount:.2f}')
    amount_display.short_description = '金额'
    
    def remarks_short(self, obj):
        """备注简短显示"""
        if obj.remarks:
            return obj.remarks[:50] + '...' if len(obj.remarks) > 50 else obj.remarks
        return '-'
    remarks_short.short_description = '备注'
    
    def reimbursement_link(self, obj):
        """显示关联的报销申请链接"""
        if obj.reimbursement:
            url = f'/admin/reimbursement/reimbursementrequest/{obj.reimbursement.id}/change/'
            return format_html('<a href="{}" target="_blank">查看报销申请</a>', url)
        return '-'
    reimbursement_link.short_description = '关联报销'
    
    def export_to_excel(self, request, queryset):
        """导出选中的记账记录到Excel"""
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "财务记账本"
        
        # 设置表头样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # 表头
        headers = ['记账日期', '提交人姓名', '报销事由', '金额', '备注']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 设置列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 40
        
        # 填充数据
        for row_num, entry in enumerate(queryset.order_by('entry_date'), 2):
            ws.cell(row=row_num, column=1, value=entry.entry_date.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row_num, column=2, value=entry.real_name)
            ws.cell(row=row_num, column=3, value=entry.reason)
            
            # 金额显示（收入为正，支出为负）
            amount_value = float(entry.amount) if entry.entry_type == 'income' else -float(entry.amount)
            amount_cell = ws.cell(row=row_num, column=4, value=amount_value)
            amount_cell.number_format = '#,##0.00'
            
            ws.cell(row=row_num, column=5, value=entry.remarks or '')
        
        # 保存到BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 返回Excel文件
        filename = f'财务记账本_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        self.message_user(request, f'成功导出 {queryset.count()} 条记账记录到Excel')
        return response
    
    export_to_excel.short_description = '📊 导出选中记录到Excel'
    
    def calculate_balance(self, request, queryset):
        """计算账户余额"""
        from decimal import Decimal
        
        # 计算所有记录（不仅是选中的）
        all_entries = AccountBook.objects.all()
        
        total_income = Decimal('0.00')
        total_expense = Decimal('0.00')
        
        for entry in all_entries:
            if entry.entry_type == 'income':
                total_income += entry.amount
            else:  # reimbursement 或 expense
                total_expense += entry.amount
        
        balance = total_income - total_expense
        
        # 显示消息
        message = (
            f'💰 账户余额计算结果：\n'
            f'总收入：¥{total_income:,.2f}\n'
            f'总支出：¥{total_expense:,.2f}\n'
            f'当前余额：¥{balance:,.2f}'
        )
        
        if balance >= 0:
            self.message_user(request, message, level='success')
        else:
            self.message_user(request, message + '\n⚠️ 警告：余额不足！', level='warning')
    
    calculate_balance.short_description = '💰 计算账户余额'

